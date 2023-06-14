import openpyxl as xl, sys, os.path, io, logging

from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from googleapiclient.http import MediaIoBaseDownload

SCOPES = ['https://www.googleapis.com/auth/spreadsheets',
              'https://www.googleapis.com/auth/drive']

def load_from_drive(file_id):
    creds = None
    # The file token.json stores the user's access and refresh tokens, and is
    # created automatically when the authorization flow completes for the first
    # time.
    if os.path.exists('test_data/token.json'):
        creds = Credentials.from_authorized_user_file(
            'test_data/token.json', SCOPES)
    # If there are no (valid) credentials available, let the user log in.
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(
                'test_data/epic_client.json', SCOPES)
            creds = flow.run_local_server(port=0)
        # Save the credentials for the next run
        with open('test_data/token.json', 'w') as token:
            token.write(creds.to_json())

    try:
        service = build('drive', 'v3', credentials=creds)
        request = service.files().export_media(fileId=file_id,
                                               mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        file = io.BytesIO()
        downloader = MediaIoBaseDownload(file, request)
        done = False
        while done is False:
            status, done = downloader.next_chunk()
            logging.info(F'Download {int(status.progress() * 100)}.')

    except HttpError as error:
        # TODO(developer) - Handle errors from drive API.
        logging.error(f'An error occurred: {error}')

    with open('aboba.xlsx', 'wb') as input:
        input.write(file.getvalue())


def read_from_sheet(file_id, range_name):
    creds = None
    # The file token.json stores the user's access and refresh tokens, and is
    # created automatically when the authorization flow completes for the first
    # time.
    if os.path.exists('test_data/token.json'):
        creds = Credentials.from_authorized_user_file('test_data/token.json', SCOPES)
    # If there are no (valid) credentials available, let the user log in.
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(
                'test_data/epic_client.json', SCOPES)
            creds = flow.run_local_server(port=0)
        # Save the credentials for the next run
        with open('test_data/token.json', 'w') as token:
            token.write(creds.to_json())

    try:
        service = build('sheets', 'v4', credentials=creds)

        # Call the Sheets API
        sheet = service.spreadsheets()
        result = sheet.values().get(spreadsheetId=file_id,
                                    range=range_name).execute()
        values = result.get('values', [])

        if not values:
            print('No data found.')
            return

        return values
    except HttpError as err:
        print(err)

if __name__ == '__main__':
    logging.basicConfig(filename='excel.log', filemode='w', level=logging.DEBUG, format='%(asctime)s | %(levelname)s | %(funcName)s, %(lineno)d: %(message)s')
    print(read_from_sheet(sys.argv[1], sys.argv[2]))
    # load_from_drive(sys.argv[1])

    # a = xl.load_workbook('aboba.xlsx')
    # sheet = a['ИУ4-23Б']

    # for row in sheet.iter_rows(min_row=3, max_row=sheet.max_row - 2):
    #     print(f'{row[0].value}: ', end='')
    #     for cell in row[1:]:
    #         if cell.value == None:
    #             print('')
    #             break
    #         print(cell.value, end=' ')
